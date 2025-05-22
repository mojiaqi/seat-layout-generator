import streamlit as st
import pandas as pd
import json
import openpyxl
import re
import io
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="座位布局生成器", layout="wide")

# 函数：将颜色转换为有效的aRGB格式
def convert_to_argb(color):
    # 如果已经是8位16进制（可能是aRGB格式）
    if re.match(r'^[0-9A-Fa-f]{8}$', color):
        return color.upper()
    
    # 如果是#开头的6位16进制（常见RGB格式）
    if re.match(r'^#[0-9A-Fa-f]{6}$', color):
        return 'FF' + color[1:].upper()
    
    # 如果是6位16进制（无#前缀）
    if re.match(r'^[0-9A-Fa-f]{6}$', color):
        return 'FF' + color.upper()
    
    # 如果是#开头的3位16进制（简写RGB格式）
    if re.match(r'^#[0-9A-Fa-f]{3}$', color):
        r, g, b = color[1], color[2], color[3]
        return 'FF' + r + r + g + g + b + b
    
    # 如果是3位16进制（无#前缀）
    if re.match(r'^[0-9A-Fa-f]{3}$', color):
        r, g, b = color[0], color[1], color[2]
        return 'FF' + r + r + g + g + b + b
    
    # 默认返回白色
    st.warning(f'警告: 无法识别的颜色格式 "{color}"，使用默认白色')
    return 'FFFFFFFF'

def process_files(excel_file, js_file, output_filename):
    try:
        # 读取JS文件内容
        js_content = js_file.read().decode('utf-8')
        
        # 解析JS内容为JSON对象
        seats_data = json.loads(js_content)
        
        # 读取Excel数据
        df_seats = pd.read_excel(excel_file, header=None)
        
        # 检查Excel文件结构
        if df_seats.shape[1] < 3:
            st.error("Excel文件格式错误：至少需要3列数据（座位号、分区、名称）")
            return None
        
        # 设置列名
        df_seats.columns = ['座位', '分区', '名称'] + [f'列{i+4}' for i in range(df_seats.shape[1]-3)]
        
        # 分析JS布局结构
        max_row = 0
        max_col = 0
        group_colors = {}

        for seat_key, seat_info in seats_data.items():
            row = seat_info.get('row', 0)
            col = seat_info.get('col', 0)
            max_row = max(max_row, row)
            max_col = max(max_col, col)
            
            group = seat_info.get('groupName')
            color = seat_info.get('color', '')
            if group and color and group not in group_colors:
                group_colors[group] = color

        # 创建Excel座位号到JS座位号的映射
        excel_to_js = {}
        for _, row in df_seats.iterrows():
            excel_seat = row['座位']
            if '-' in str(excel_seat):
                parts = str(excel_seat).split('-')
                js_seat = f'{parts[0]}_{parts[1]}'
                excel_to_js[excel_seat] = js_seat

        # 创建新的Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # 设置列宽
        for col in range(1, max_col + 2):  # +1 for row labels, +1 for extra margin
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15

        # 添加表头
        ws.cell(row=1, column=1).value = '座位布局'
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(max_col + 1)}1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 添加行标签
        for row in range(1, max_row + 1):
            ws.cell(row=row + 2, column=1).value = f'第{row}排'
            ws.cell(row=row + 2, column=1).font = Font(bold=True)
            ws.cell(row=row + 2, column=1).alignment = Alignment(horizontal='center')

        # 添加列标签
        for col in range(1, max_col + 1):
            ws.cell(row=2, column=col + 1).value = col
            ws.cell(row=2, column=col + 1).font = Font(bold=True)
            ws.cell(row=2, column=col + 1).alignment = Alignment(horizontal='center')

        # 填充座位数据
        filled_count = 0
        for seat_key, seat_info in seats_data.items():
            row = seat_info.get('row', 0)
            col = seat_info.get('col', 0)
            group = seat_info.get('groupName', '')
            color = seat_info.get('color', '')
            
            # Excel行列索引（+2是因为有表头和标签行）
            excel_row = row + 2
            excel_col = col + 1
            
            # 设置单元格值和样式
            cell = ws.cell(row=excel_row, column=excel_col)
            
            # 检查是否有对应的Excel数据
            js_seat = f'{row}_{col}'
            excel_seat = None
            for es, js in excel_to_js.items():
                if js == js_seat:
                    excel_seat = es
                    break
            
            if excel_seat:
                # 有对应的Excel数据，填充学校名称
                excel_row_data = df_seats[df_seats['座位'] == excel_seat].iloc[0]
                school_name = excel_row_data['名称']
                partition = excel_row_data['分区']
                
                cell.value = school_name
                filled_count += 1
                
                # 根据分区设置颜色
                if partition in group_colors:
                    fill_color = group_colors[partition]
                else:
                    # 如果Excel中的分区在JS中找不到对应颜色，使用JS中的原始颜色
                    fill_color = color
                
                # 转换为标准aRGB格式
                fill_color = convert_to_argb(fill_color)
                
                try:
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    cell.fill = fill
                except ValueError as e:
                    st.warning(f'颜色转换错误: {fill_color} - {e}')
                    # 使用默认颜色
                    fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                    cell.fill = fill
            else:
                # 没有对应的Excel数据，只设置颜色
                if color:
                    # 转换为标准aRGB格式
                    color = convert_to_argb(color)
                    
                    try:
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.fill = fill
                    except ValueError as e:
                        st.warning(f'颜色转换错误: {color} - {e}')
                        # 使用默认颜色
                        fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                        cell.fill = fill
            
            # 设置边框和对齐方式
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 创建颜色标签工作表
        ws_legend = wb.create_sheet(title='颜色标签')

        # 设置列宽
        ws_legend.column_dimensions['A'].width = 15
        ws_legend.column_dimensions['B'].width = 20
        ws_legend.column_dimensions['C'].width = 15

        # 添加标题
        ws_legend.cell(row=1, column=1).value = '颜色标签说明'
        ws_legend.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_legend.merge_cells('A1:C1')
        ws_legend.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # 添加表头
        headers = ['颜色', '分区', '数据量']
        for col, header in enumerate(headers, 1):
            ws_legend.cell(row=3, column=col).value = header
            ws_legend.cell(row=3, column=col).font = Font(bold=True)
            ws_legend.cell(row=3, column=col).alignment = Alignment(horizontal='center')
            ws_legend.cell(row=3, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        # 添加分区颜色示例
        row = 4
        for group, color in sorted(group_colors.items()):
            # 颜色示例 - 转换为标准aRGB格式
            color = convert_to_argb(color)
            
            try:
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws_legend.cell(row=row, column=1).fill = fill
            except ValueError as e:
                st.warning(f'图例颜色转换错误: {color} - {e}')
                # 使用默认颜色
                fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                ws_legend.cell(row=row, column=1).fill = fill
            
            # 分区名称
            ws_legend.cell(row=row, column=2).value = group
            ws_legend.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 数据量（计算该分区在JS中的座位数量）
            count = sum(1 for seat in seats_data.values() if seat.get('groupName') == group)
            ws_legend.cell(row=row, column=3).value = count
            ws_legend.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            # 添加边框
            for col in range(1, 4):
                ws_legend.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            row += 1

        # 添加Excel数据分区统计
        row += 2
        ws_legend.cell(row=row, column=1).value = 'Excel数据分区统计'
        ws_legend.cell(row=row, column=1).font = Font(bold=True)
        ws_legend.merge_cells(f'A{row}:C{row}')
        ws_legend.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        row += 1
        headers = ['分区', '数据量', '']
        for col, header in enumerate(headers, 1):
            ws_legend.cell(row=row, column=col).value = header
            ws_legend.cell(row=row, column=col).font = Font(bold=True)
            ws_legend.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            ws_legend.cell(row=row, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        row += 1
        for partition, count in df_seats['分区'].value_counts().items():
            ws_legend.cell(row=row, column=1).value = partition
            ws_legend.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            ws_legend.cell(row=row, column=2).value = count
            ws_legend.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 添加边框
            for col in range(1, 4):
                ws_legend.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            row += 1

        # 在主工作表添加提示
        ws.cell(row=max_row + 4, column=1).value = '注: 请查看「颜色标签」工作表了解颜色与分区的对应关系'
        ws.cell(row=max_row + 4, column=1).font = Font(italic=True)
        ws.merge_cells(f'A{max_row + 4}:{get_column_letter(max_col + 1)}{max_row + 4}')

        # 保存到内存中
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'file': excel_buffer,
            'filename': output_filename,
            'stats': {
                'total_seats': len(seats_data),
                'filled_seats': filled_count,
                'excel_rows': len(df_seats),
                'max_row': max_row,
                'max_col': max_col,
                'partitions': len(group_colors)
            }
        }
    
    except Exception as e:
        st.error(f"处理文件时出错: {str(e)}")
        return None

# 应用标题
st.title("座位布局生成器")
st.markdown("上传Excel座位数据和JS布局文件，自动生成带颜色标签的座位布局Excel文件。")

# 文件上传区域
with st.container():
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("上传Excel座位数据")
        excel_file = st.file_uploader("选择Excel文件", type=["xlsx", "xls"], help="Excel文件应包含座位号、分区和名称三列数据")
        
    with col2:
        st.subheader("上传JS布局文件")
        js_file = st.file_uploader("选择JS文件", type=["js"], help="JS文件应包含座位布局的JSON数据")

# 输出文件名设置
output_filename = st.text_input("输出文件名", "座位布局填充结果.xlsx")

# 处理按钮
if st.button("生成座位布局"):
    if excel_file is None or js_file is None:
        st.warning("请上传Excel座位数据和JS布局文件")
    else:
        with st.spinner("正在处理文件..."):
            result = process_files(excel_file, js_file, output_filename)
            
            if result:
                # 显示处理统计信息
                st.success("座位布局生成成功！")
                
                stats = result['stats']
                st.subheader("处理统计")
                col1, col2, col3 = st.columns(3)
                col1.metric("总座位数", stats['total_seats'])
                col1.metric("填充座位数", stats['filled_seats'])
                col2.metric("Excel数据行数", stats['excel_rows'])
                col2.metric("分区数量", stats['partitions'])
                col3.metric("布局行数", stats['max_row'])
                col3.metric("布局列数", stats['max_col'])
                
                # 提供下载链接
                st.download_button(
                    label="下载座位布局Excel文件",
                    data=result['file'],
                    file_name=result['filename'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# 使用说明
with st.expander("使用说明"):
    st.markdown("""
    ### 文件格式要求
    
    #### Excel座位数据文件
    - 第一列：座位号（格式为"行-列"，如"1-5"）
    - 第二列：分区名称
    - 第三列：学校或单位名称
    
    #### JS布局文件
    - 必须是有效的JSON格式
    - 每个座位对象应包含row、col、groupName和color属性
    
    ### 处理流程
    1. 上传Excel座位数据和JS布局文件
    2. 点击"生成座位布局"按钮
    3. 系统自动将Excel数据映射到JS布局中
    4. 生成带有颜色标签的Excel文件
    5. 下载生成的Excel文件
    
    ### 注意事项
    - 座位号格式必须匹配，Excel中的"行-列"会自动转换为JS中的"行_列"
    - 分区颜色会根据JS文件中的设置自动应用
    - 生成的Excel文件包含主表和颜色标签表两个工作表
    """)

# 页脚
st.markdown("---")
st.markdown("座位布局生成器 © 2025")
